from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment

def apply_color(cell, value):
    if value == 1:
        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    elif value > 0.75:
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    elif value > 0.5:
        cell.fill = PatternFill(start_color="E0FFD1", end_color="E0FFD1", fill_type="solid")
    elif value < -0.5:
        cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
    elif value < -0.25:
        cell.fill = PatternFill(start_color="FFD1D1", end_color="FFD1D1", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def save_to_workbook(correlation_matrix, starting_date):
    wb = Workbook()
    ws = wb.active
    ws.title = "Correlation Matrix"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=correlation_matrix.shape[1] + 1)
    ws['A1'] = f'Correlation Matrix (since {starting_date})'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    for r_idx, row in enumerate(dataframe_to_rows(correlation_matrix, index=True, header=True), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if r_idx > 2 and c_idx > 1 and isinstance(value, (int, float)):
                apply_color(cell, value)

    wb.save("correlation_matrix.xlsx")
